"""Parse an uploaded deployment-form ``.xlsx`` back into raw field values.

The parser is intentionally template-agnostic: it reads the workbook's own hidden
``__meta__`` sheet (written by :mod:`deployment_form_excel`) to learn where each
``field_key`` lives, then reads those cells. It returns raw ``{field_key: value}``
plus the metadata block. Mapping raw values into the wizard ``answers`` object is
done by the service against the *current* template, so stale/locked keys are
dropped and the live template stays the source of truth.
"""

from __future__ import annotations

import json
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from .deployment_form_excel import MAIN_SHEET, META_SHEET


def _maybe_json(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if text and text[0] in "[{":
        try:
            return json.loads(text)
        except (ValueError, TypeError):
            return value
    return value


def parse_form(file_bytes: bytes) -> Tuple[Dict[str, Any], Dict[str, Any], List[str]]:
    """Read a workbook's metadata + filled values.

    Returns ``(raw_values, metadata, errors)`` where ``raw_values`` maps each
    editable ``field_key`` to its cell value. ``errors`` is a list of blocking
    parse problems (e.g. wrong file, missing metadata).
    """
    errors: List[str] = []
    try:
        # Normal (not read_only) mode so we can address value cells by coordinate.
        wb = load_workbook(io_bytes(file_bytes), data_only=True)
    except Exception:  # openpyxl raises a variety of types on bad input
        return {}, {}, ["The uploaded file is not a valid .xlsx workbook."]

    if META_SHEET not in wb.sheetnames or MAIN_SHEET not in wb.sheetnames:
        return {}, {}, ["This file is not a Kubesight deployment form (missing metadata)."]

    meta_ws = wb[META_SHEET]
    metadata: Dict[str, Any] = {}
    for row in meta_ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if not row or row[0] is None:
            continue
        metadata[str(row[0])] = _maybe_json(row[1])

    field_map = metadata.get("fieldMap")
    if not isinstance(field_map, dict) or not field_map:
        return {}, metadata, ["This deployment form is missing its field map and cannot be parsed."]

    main_ws = wb[MAIN_SHEET]
    raw_values: Dict[str, Any] = {}
    for key, spec in field_map.items():
        if not isinstance(spec, dict):
            continue
        if spec.get("locked"):
            # Locked values are always re-sourced from the live template.
            continue
        coord = spec.get("cell")
        if not coord:
            continue
        try:
            value = main_ws[coord].value
        except (ValueError, KeyError):
            continue
        if value is not None and str(value).strip() != "":
            raw_values[key] = value

    return raw_values, metadata, errors


def io_bytes(file_bytes: bytes):
    """openpyxl accepts a file-like object; wrap the bytes."""
    import io

    return io.BytesIO(file_bytes)


def read_metadata(file_bytes: bytes) -> Optional[Dict[str, Any]]:
    """Read just the metadata block (used for lightweight forgery/expiry checks)."""
    _, metadata, errors = parse_form(file_bytes)
    if errors:
        return None
    return metadata
