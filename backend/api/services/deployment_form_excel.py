"""Generate the fillable ``.xlsx`` deployment form from a field list.

Layout:
  * **Deployment Request** — the visible sheet the user fills. Grouped sections
    (Basic Info, Image Version, ...) with Label / Value / Notes columns. Locked
    rows are greyed and protected; editable rows are unlocked; dropdown rows carry
    a list ``DataValidation`` referencing the hidden ``__lists__`` sheet.
  * **__lists__** (hidden) — one column per dropdown, holding its allowed values,
    so validations reference a range (no 255-char inline-list limit).
  * **__meta__** (hidden) — template id/version, schema version, form id, generator
    identity, and a ``field_key -> {cell,type,locked}`` map used for
    position-independent parsing on import.

Sheet protection is a UX guard only — the server re-validates every value on
import and never trusts locked cells from the file.
"""

from __future__ import annotations

import io
import json
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .deployment_form_schema import GROUP_ORDER

MAIN_SHEET = "Deployment Request"
LISTS_SHEET = "__lists__"
META_SHEET = "__meta__"

# Styling
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_GROUP_FILL = PatternFill("solid", fgColor="374151")
_LOCKED_FILL = PatternFill("solid", fgColor="E5E7EB")
_INPUT_FILL = PatternFill("solid", fgColor="FFFBEB")
_WHITE = Font(color="FFFFFF", bold=True)
_GROUP_FONT = Font(color="FFFFFF", bold=True, size=12)
_LABEL_FONT = Font(bold=True)
_REQUIRED_FONT = Font(bold=True, color="B91C1C")
_HELP_FONT = Font(italic=True, color="6B7280", size=9)
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _dropdown_values(field: Dict[str, Any], dropdown_data: Dict[str, List[str]]) -> List[str]:
    if field.get("options"):
        return [str(v) for v in field["options"]]
    key = field.get("dropdown")
    if key:
        return [str(v) for v in (dropdown_data.get(key) or [])]
    return []


def build_workbook(
    fields: List[Dict[str, Any]],
    dropdown_data: Dict[str, List[str]],
    metadata: Dict[str, Any],
) -> bytes:
    """Render the workbook and return its bytes."""
    wb = Workbook()
    main = wb.active
    main.title = MAIN_SHEET
    lists = wb.create_sheet(LISTS_SHEET)
    meta = wb.create_sheet(META_SHEET)

    # ---- DataValidation (dropdown) factory ----
    # Prefer an *inline* comma-separated list ("a,b,c") — the most compatible form
    # that always renders the dropdown arrow in Excel. Fall back to a hidden
    # __lists__ column only for long option sets that exceed Excel's ~255-char
    # inline limit (e.g. many namespaces). Cross-sheet references are the usual
    # reason dropdowns silently vanish, so we avoid them for the common case.
    dv_cache: Dict[str, DataValidation] = {}
    lists_state = {"col": 0}

    def dropdown_for(values: List[str]) -> DataValidation:
        signature = json.dumps(values)
        cached = dv_cache.get(signature)
        if cached is not None:
            return cached
        joined = ",".join(values)
        inline_safe = all("," not in v and '"' not in v for v in values) and len(joined) <= 250
        if inline_safe:
            dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
        else:
            lists_state["col"] += 1
            col_letter = get_column_letter(lists_state["col"])
            lists.cell(row=1, column=lists_state["col"], value=f"list{lists_state['col']}")
            for i, val in enumerate(values, start=2):
                lists.cell(row=i, column=lists_state["col"], value=val)
            end = len(values) + 1
            dv = DataValidation(
                type="list",
                formula1=f"'{LISTS_SHEET}'!${col_letter}$2:${col_letter}${end}",
                allow_blank=True,
            )
        # Warn-only: still offer the dropdown, but let users type a value not in the
        # list (e.g. a namespace that wasn't listed) — the server re-validates.
        dv.showErrorMessage = False
        main.add_data_validation(dv)
        dv_cache[signature] = dv
        return dv

    lists.sheet_state = "hidden"

    # ---- Main sheet ----
    main.column_dimensions["A"].width = 34
    main.column_dimensions["B"].width = 40
    main.column_dimensions["C"].width = 52

    # Title banner
    main.merge_cells("A1:C1")
    title = main.cell(row=1, column=1, value=f"Deployment Request — {metadata.get('templateName') or metadata.get('templateId')}")
    title.font = Font(color="FFFFFF", bold=True, size=14)
    title.fill = _HEADER_FILL
    title.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    main.row_dimensions[1].height = 26
    main.merge_cells("A2:C2")
    subtitle = main.cell(
        row=2, column=1,
        value="Fill the highlighted (cream) cells. Grey rows are set by the template — you "
              "don't need to change them. Do not edit the hidden sheets. Upload this file "
              "back into Kubesight.",
    )
    subtitle.font = _HELP_FONT
    subtitle.alignment = Alignment(vertical="center", horizontal="left", indent=1)

    # Column headers
    row = 3
    for col, header in enumerate(("Field", "Value", "Notes"), start=1):
        cell = main.cell(row=row, column=col, value=header)
        cell.font = _WHITE
        cell.fill = _GROUP_FILL
        cell.border = _BORDER

    field_map: Dict[str, Any] = {}
    # Group and preserve GROUP_ORDER; within a group preserve field order.
    by_group: Dict[str, List[Dict[str, Any]]] = {}
    for field in fields:
        by_group.setdefault(field["group"], []).append(field)

    dv_by_range: Dict[str, DataValidation] = {}

    row = 4
    for group in GROUP_ORDER:
        group_fields = by_group.get(group)
        if not group_fields:
            continue
        main.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        gcell = main.cell(row=row, column=1, value=group)
        gcell.font = _GROUP_FONT
        gcell.fill = _GROUP_FILL
        gcell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        main.row_dimensions[row].height = 20
        row += 1

        for field in group_fields:
            label = field["label"] + (" *" if field.get("required") else "")
            label_cell = main.cell(row=row, column=1, value=label)
            label_cell.font = _REQUIRED_FONT if field.get("required") else _LABEL_FONT
            label_cell.border = _BORDER
            label_cell.alignment = Alignment(vertical="center", wrap_text=True)

            value_cell = main.cell(row=row, column=2, value=field.get("default") or "")
            value_cell.border = _BORDER
            value_cell.alignment = Alignment(vertical="center")

            help_cell = main.cell(row=row, column=3, value=field.get("help") or "")
            help_cell.font = _HELP_FONT
            help_cell.border = _BORDER
            help_cell.alignment = Alignment(vertical="center", wrap_text=True)

            coord = value_cell.coordinate
            locked = bool(field.get("locked"))
            if locked:
                # Grey shading marks template-controlled fields. We do NOT protect
                # the sheet: Excel sheet-protection blocks editing in ways that
                # confuse users, and it buys nothing here — on import every locked
                # field is re-sourced from the live template and the file's value
                # for it is ignored. The shading is guidance, the server is the gate.
                label_cell.fill = _LOCKED_FILL
                value_cell.fill = _LOCKED_FILL
            else:
                value_cell.fill = _INPUT_FILL
                values = _dropdown_values(field, dropdown_data)
                if values:
                    dropdown_for([str(v) for v in values]).add(value_cell)

            field_map[field["key"]] = {
                "cell": coord,
                "type": field.get("type") or "text",
                "locked": locked,
            }
            row += 1

    main.freeze_panes = "A4"

    # ---- Hidden meta sheet ----
    meta.sheet_state = "hidden"
    full_meta = dict(metadata)
    full_meta["fieldMap"] = field_map
    allowed = [k for k, v in field_map.items() if not v["locked"]]
    locked_keys = [k for k, v in field_map.items() if v["locked"]]
    required_keys = [f["key"] for f in fields if f.get("required")]
    full_meta["allowedFields"] = allowed
    full_meta["lockedFields"] = locked_keys
    full_meta["requiredFields"] = required_keys

    meta.cell(row=1, column=1, value="key")
    meta.cell(row=1, column=2, value="value")
    r = 2
    # Store each metadata entry as key + JSON value so parsing is trivial.
    for key, value in full_meta.items():
        meta.cell(row=r, column=1, value=key)
        meta.cell(row=r, column=2, value=value if isinstance(value, str) else json.dumps(value))
        r += 1
    meta.protection.sheet = True

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
