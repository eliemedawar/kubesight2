"""Tests for the Smart Deployment Form: schema, Excel round-trip, validation, API."""

import io

from openpyxl import load_workbook

from api.services.deployment_form_excel import MAIN_SHEET, build_workbook
from api.services.deployment_form_parser import parse_form
from api.services.deployment_form_schema import (
    assemble_answers,
    build_form_fields,
    template_version,
)
from api.services.deployment_form_validation import validate_import
from api.services.template_resolver import resolve_template
from tests.conftest import auth_headers


def _template(**schema):
    return {
        "id": "orders-api",
        "name": "Orders API",
        "workloadType": "Deployment",
        "containers": [{"name": "app", "image": "acme/orders", "tag": "1.0.0", "ports": [8080], "pullPolicy": "Always"}],
        "resources": {"cpuRequest": "100m", "cpuLimit": "500m", "memoryRequest": "128Mi", "memoryLimit": "256Mi"},
        "networking": {
            "service": {"enabled": True, "type": "ClusterIP", "port": 8080, "targetPort": 8080},
            "ingress": {"enabled": False, "host": "orders.example.com", "path": "/"},
        },
        "scaling": {"replicas": 2, "hpa": {"enabled": False, "minReplicas": 1, "maxReplicas": 5}},
        "schema": schema,
    }


def _fill(fields, values):
    """Return a copy of the fields with defaults set from ``values`` (by key)."""
    for field in fields:
        if field["key"] in values:
            field["default"] = values[field["key"]]
    return fields


# ---------------------------------------------------------------------------
# Field schema
# ---------------------------------------------------------------------------

def test_locked_fields_not_editable_and_overrides_exposed():
    fields = build_form_fields(_template(overrides={"tag": True, "serviceType": ["ClusterIP", "NodePort"]}))
    by_key = {f["key"]: f for f in fields}

    # Workload kind is a locked, display-only row.
    assert by_key["locked.workloadType"]["locked"] is True
    assert by_key["locked.workloadType"]["answerPath"] is None

    # Tag override is editable and required, service type is a dropdown of the allowed set.
    assert by_key["overrides.tag"]["locked"] is False
    assert by_key["overrides.serviceType"]["options"] == ["ClusterIP", "NodePort"]
    # Image is locked because the schema does not allow overriding it.
    assert "overrides.image" not in by_key
    assert "locked.image" in by_key


def test_env_is_create_only_no_existing_pickers():
    fields = build_form_fields(
        _template(env=[{"key": "DB_PASSWORD", "required": True, "sensitive": True}])
    )
    by_key = {f["key"]: f for f in fields}
    # Create-only: just a value cell (+ data key). No existing pickers, no source enum.
    assert "env.DB_PASSWORD.source" not in by_key
    assert "env.DB_PASSWORD.existingSecret" not in by_key
    assert "env.DB_PASSWORD.existingConfigMap" not in by_key
    value_field = by_key["env.DB_PASSWORD.value"]
    assert value_field["sensitive"] is True
    assert value_field["default"] == ""


def test_data_key_only_shown_when_value_becomes_configmap_or_secret():
    by_key = lambda fs: {f["key"]: f for f in fs}  # noqa: E731
    # Plain-value-only var → no data key cell (it never lands in a ConfigMap/Secret).
    plain = by_key(build_form_fields(_template(env=[{"key": "LOG_LEVEL", "allowedSources": ["value"]}])))
    assert "env.LOG_LEVEL.value" in plain
    assert "env.LOG_LEVEL.key" not in plain
    # A var that can create a ConfigMap → data key cell present.
    cm = by_key(build_form_fields(_template(env=[{"key": "CFG", "allowedSources": ["value", "createConfigMap"]}])))
    assert "env.CFG.key" in cm


def test_env_value_creates_configmap_or_secret():
    """A filled value creates a ConfigMap (non-sensitive) or Secret (sensitive);
    a blank value falls back to the template default (returns None)."""
    from api.services.deployment_form_schema import _infer_env_entry

    cm_meta = {"envAllowed": ["value", "createConfigMap"], "envSensitive": False,
               "createConfigMapName": "app-config", "createSecretName": "app-secret"}
    assert _infer_env_entry({"value": "INFO"}, cm_meta) == {
        "source": "createConfigMap", "configMapName": "app-config", "value": "INFO",
    }
    # Sensitive → creates a Secret with the autofilled name.
    sec_meta = {"envAllowed": ["existingSecret", "createSecret"], "envSensitive": True,
                "createSecretName": "app-secret"}
    assert _infer_env_entry({"value": "s3cr3t"}, sec_meta) == {
        "source": "createSecret", "secretName": "app-secret", "value": "s3cr3t",
    }
    # Template that only permits a literal value → literal.
    assert _infer_env_entry({"value": "x"}, {"envAllowed": ["value"], "envSensitive": False}) == {
        "source": "value", "value": "x",
    }
    # Blank value → use the template default.
    assert _infer_env_entry({}, cm_meta) is None


# ---------------------------------------------------------------------------
# Excel round-trip: generate -> parse -> assemble
# ---------------------------------------------------------------------------

def test_excel_roundtrip_reconstructs_answers():
    tpl = _template(
        overrides={"tag": True, "replicas": True},
        env=[{"key": "LOG_LEVEL", "required": False, "default": "INFO"}],
    )
    fields = build_form_fields(tpl)
    _fill(fields, {
        "basics.clusterId": "c1",
        "basics.namespace": "shop",
        "basics.appName": "orders",
        "overrides.tag": "2.1.0",
        "overrides.replicas": "3",
        "env.LOG_LEVEL.value": "DEBUG",
        "changeSummary": "ship it",
    })
    xlsx = build_workbook(fields, {"clusters": ["c1"], "namespaces": ["shop"]}, {"templateId": "orders-api"})

    raw_values, metadata, errors = parse_form(xlsx)
    assert errors == []
    assert metadata["templateId"] == "orders-api"

    # Rebuild answers against the current template (mirrors the import service).
    answers = assemble_answers(raw_values, build_form_fields(tpl))
    assert answers["basics"] == {"clusterId": "c1", "namespace": "shop", "appName": "orders"}
    assert answers["overrides"]["tag"] == "2.1.0"
    assert answers["overrides"]["replicas"] == 3  # coerced to int
    # Create-only: the value creates a ConfigMap with the autofilled name.
    assert answers["env"]["LOG_LEVEL"] == {
        "source": "createConfigMap", "configMapName": "orders-api-config", "value": "DEBUG",
    }
    assert answers["changeSummary"] == "ship it"
    # Locked workload kind never round-trips into answers.
    assert "workloadType" not in answers

    # The reconstructed answers resolve cleanly against the template.
    payload, err = resolve_template(tpl, answers)
    assert err is None
    assert payload["containers"][0]["tag"] == "2.1.0"
    assert payload["scaling"]["replicas"] == 3


def test_workbook_is_editable_with_inline_dropdowns():
    """The sheet must not be protected (users edit freely) and dropdowns must use
    inline lists that render reliably in Excel."""
    tpl = _template(overrides={"tag": True, "serviceType": ["ClusterIP", "NodePort"]})
    fields = build_form_fields(tpl)
    xlsx = build_workbook(fields, {"clusters": ["c1"], "namespaces": ["a", "b"]}, {"templateId": "orders-api"})
    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(xlsx))[MAIN_SHEET]
    assert ws.protection.sheet is False
    dvs = ws.data_validations.dataValidation
    assert dvs, "expected dropdowns"
    # Service type + namespaces render as inline lists (quoted), not sheet refs.
    formulas = [dv.formula1 for dv in dvs]
    assert any('"ClusterIP,NodePort"' == f for f in formulas)
    assert any("a,b" in f for f in formulas)


def test_create_secret_value_roundtrips_and_provisions():
    """A sensitive var with source=createSecret + a value must round-trip and make
    the resolver provision a Secret (wizard parity)."""
    tpl = _template(env=[{"key": "DB_PASSWORD", "required": True, "sensitive": True}])
    fields = build_form_fields(tpl)
    # Just fill the value — the source (createSecret) and the Secret name are inferred.
    _fill(fields, {
        "basics.clusterId": "c1", "basics.namespace": "shop", "basics.appName": "orders",
        "env.DB_PASSWORD.value": "s3cr3t",
    })
    xlsx = build_workbook(fields, {}, {"templateId": "orders-api"})
    raw, _, errors = parse_form(xlsx)
    assert errors == []
    answers = assemble_answers(raw, build_form_fields(tpl))
    assert answers["env"]["DB_PASSWORD"]["source"] == "createSecret"
    assert answers["env"]["DB_PASSWORD"]["value"] == "s3cr3t"
    # Name autofilled from the template name (Orders API -> orders-api-secret).
    assert answers["env"]["DB_PASSWORD"]["secretName"] == "orders-api-secret"

    payload, err = resolve_template(tpl, answers)
    assert err is None
    provisioned = payload["environment"]["provisionedSecrets"]
    assert any(s["name"] == "orders-api-secret" for s in provisioned)


def test_volume_mounts_excluded_from_form_and_validation_warns():
    """Volume mounts aren't in the Excel — they're completed in the wizard. The form
    has no volume cells, and validation warns (not blocks) instead of demanding a name."""
    tpl = _template(volumeMounts=[{"mountPath": "/opt/certs", "kind": "configMap"}])
    fields = build_form_fields(tpl)
    assert not any(f["key"].startswith("volumes.") for f in fields)

    answers = {"basics": {"appName": "orders", "namespace": "shop", "clusterId": "c1"}}
    result = validate_import(
        tpl, answers,
        dropdown_data={"namespaces": ["shop"]},
        context={"clusterAccessible": True, "namespaceAccessible": True, "namespace": "shop",
                 "templateExists": True, "versionMatches": True},
    )
    assert result["blocking"] is False
    assert any(c["field"] == "volumes" and c["level"] == "warn" for c in result["checks"])


def test_parse_rejects_non_form_workbook():
    from openpyxl import Workbook

    wb = Workbook()
    wb.active["A1"] = "not a form"
    buf = io.BytesIO()
    wb.save(buf)
    _, _, errors = parse_form(buf.getvalue())
    assert errors and "not a Kubesight deployment form" in errors[0]


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def test_validation_blocks_when_namespace_access_denied():
    tpl = _template()
    answers = {"basics": {"appName": "orders", "namespace": "prod", "clusterId": "c1"}}
    result = validate_import(
        tpl, answers,
        dropdown_data={"namespaces": ["prod"]},
        context={"clusterAccessible": True, "namespaceAccessible": False, "namespace": "prod",
                 "templateExists": True, "versionMatches": True},
    )
    assert result["blocking"] is True
    assert any(c["field"] == "basics.namespace" and c["level"] == "error" for c in result["checks"])


def test_validation_flags_missing_secret_reference():
    tpl = _template(env=[{"key": "API_KEY", "sensitive": True, "allowedSources": ["existingSecret"]}])
    answers = {
        "basics": {"appName": "orders", "namespace": "prod", "clusterId": "c1"},
        "env": {"API_KEY": {"source": "existingSecret", "secretName": "missing-secret"}},
    }
    result = validate_import(
        tpl, answers,
        dropdown_data={"namespaces": ["prod"], "secrets": ["other-secret"]},
        context={"clusterAccessible": True, "namespaceAccessible": True, "namespace": "prod",
                 "templateExists": True, "versionMatches": True},
    )
    assert any(c["level"] == "error" and "secret" in c["message"].lower() for c in result["checks"])
    assert result["blocking"] is True


def test_validation_passes_for_clean_import():
    tpl = _template(overrides={"tag": True})
    answers = {"basics": {"appName": "orders", "namespace": "shop", "clusterId": "c1"}, "overrides": {"tag": "2.0"}}
    result = validate_import(
        tpl, answers,
        dropdown_data={"namespaces": ["shop"]},
        context={"clusterAccessible": True, "namespaceAccessible": True, "namespace": "shop",
                 "templateExists": True, "versionMatches": True},
    )
    assert result["blocking"] is False
    assert any(c["level"] == "ok" for c in result["checks"])


def test_version_drift_warns():
    tpl = _template()
    assert template_version(tpl) != template_version(_template(overrides={"tag": True}))
    result = validate_import(
        tpl, {"basics": {"appName": "orders", "namespace": "shop", "clusterId": "c1"}},
        dropdown_data={}, context={"templateExists": True, "versionMatches": False},
    )
    assert any(c["level"] == "warn" and "template" in c["field"] for c in result["checks"])


# ---------------------------------------------------------------------------
# API: generate -> import -> apply-to-wizard
# ---------------------------------------------------------------------------

def test_dropdowns_mirror_wizard_mock_data(app):
    """Namespaces/storage classes must fall back to the same mock data the wizard
    uses, so the Excel gets real dropdowns even on mock/demo clusters."""
    from api.services.deployment_form_service import collect_form_dropdowns

    dd = collect_form_dropdowns(None, "prod-us-east", "payments")
    assert "namespaces" in dd and "payments" in dd["namespaces"]
    assert dd.get("storageClasses")  # gp3/efs from mock data


def _create_template(client, token):
    resp = client.post(
        "/api/inventory/deploy/wizard/templates",
        headers=auth_headers(token),
        json={
            "name": "Orders API",
            "category": "Backend",
            "workloadType": "Deployment",
            "containers": [{"name": "app", "image": "acme/orders", "tag": "1.0.0", "ports": [8080]}],
            "resources": {"cpuRequest": "100m", "memoryRequest": "128Mi"},
            "schema": {"overrides": {"tag": True}, "env": [{"key": "LOG_LEVEL", "default": "INFO"}]},
        },
    )
    assert resp.status_code == 201, resp.get_data(as_text=True)
    return resp.get_json()["data"]["id"]


def test_generate_import_apply_flow(client, admin_token):
    template_id = _create_template(client, admin_token)

    gen = client.post(
        "/api/deployment-forms/generate",
        headers=auth_headers(admin_token),
        json={"templateId": template_id},
    )
    assert gen.status_code == 200, gen.get_data(as_text=True)
    assert gen.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    xlsx = gen.get_data()

    # Fill required fields (namespace + appName) in the generated workbook.
    wb = load_workbook(io.BytesIO(xlsx))
    meta = {row[0]: row[1] for row in wb["__meta__"].iter_rows(min_row=2, max_col=2, values_only=True)}
    import json as _json
    field_map = _json.loads(meta["fieldMap"])
    ws = wb[MAIN_SHEET]
    ws[field_map["basics.clusterId"]["cell"]] = "demo-cluster"
    ws[field_map["basics.namespace"]["cell"]] = "default"
    ws[field_map["basics.appName"]["cell"]] = "orders"
    ws[field_map["overrides.tag"]["cell"]] = "2.3.4"
    buf = io.BytesIO()
    wb.save(buf)

    imp = client.post(
        "/api/deployment-forms/import",
        headers=auth_headers(admin_token),
        data={"file": (io.BytesIO(buf.getvalue()), "form.xlsx")},
        content_type="multipart/form-data",
    )
    assert imp.status_code == 200, imp.get_data(as_text=True)
    body = imp.get_json()["data"]
    import_id = body["id"]
    assert body["answers"]["overrides"]["tag"] == "2.3.4"
    assert body["answers"]["basics"]["namespace"] == "default"

    applied = client.post(
        f"/api/deployment-forms/imports/{import_id}/apply-to-wizard",
        headers=auth_headers(admin_token),
    )
    assert applied.status_code == 200, applied.get_data(as_text=True)
    data = applied.get_json()["data"]
    assert data["template"]["id"] == template_id
    assert data["answers"]["basics"]["appName"] == "orders"

    # Add-to-bundle reuses the existing change-bundle workflow.
    bundled = client.post(
        f"/api/deployment-forms/imports/{import_id}/add-to-bundle",
        headers=auth_headers(admin_token),
    )
    assert bundled.status_code == 200, bundled.get_data(as_text=True)
    bundle = bundled.get_json()["data"]["bundle"]
    assert bundle["items"], "expected one staged change in the bundle"
    assert bundle["items"][0]["actionType"] == "create_from_template"
